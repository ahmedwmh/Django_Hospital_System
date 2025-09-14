from celery import shared_task
from django.conf import settings
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.utils import timezone
import os
import json
from datetime import datetime, timedelta

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel generation is optional - will work without pandas/openpyxl
PANDAS_AVAILABLE = False
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    # Create dummy classes for when pandas is not available
    class pd:
        @staticmethod
        def DataFrame(*args, **kwargs):
            return None
    class Workbook:
        def __init__(self, *args, **kwargs):
            pass

from .models import Report
from apps.patients.models import Patient, Test, Treatment, Surgery
from apps.hospital.models import City, Disease


@shared_task
def generate_patient_record_pdf(report_id):
    """Generate patient record PDF report"""
    try:
        report = Report.objects.get(id=report_id)
        report.status = 'GENERATING'
        report.save()
        
        patient_id = report.parameters.get('patient_id')
        patient = Patient.objects.select_related(
            'user', 'doctor__user', 'doctor__center__city'
        ).prefetch_related(
            'patient_diseases__disease', 'tests', 'treatments', 'surgeries'
        ).get(id=patient_id)
        
        # Create PDF
        filename = f"patient_record_{patient.patient_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("Patient Medical Record", title_style))
        story.append(Spacer(1, 20))
        
        # Patient Information
        patient_data = [
            ['Patient ID:', patient.patient_id],
            ['Name:', patient.user.get_full_name()],
            ['Email:', patient.user.email],
            ['Phone:', patient.user.phone_number],
            ['Date of Birth:', patient.date_of_birth.strftime('%Y-%m-%d')],
            ['Age:', str(patient.age)],
            ['Gender:', patient.get_gender_display()],
            ['Blood Group:', patient.blood_group or 'Not specified'],
            ['Address:', patient.address],
            ['Emergency Contact:', f"{patient.emergency_contact_name} ({patient.emergency_contact_phone})"],
            ['Doctor:', patient.doctor.user.get_full_name()],
            ['Center:', patient.doctor.center.name],
            ['City:', patient.doctor.center.city.name],
        ]
        
        patient_table = Table(patient_data, colWidths=[2*inch, 4*inch])
        patient_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ]))
        
        story.append(Paragraph("Patient Information", styles['Heading2']))
        story.append(patient_table)
        story.append(Spacer(1, 20))
        
        # Medical History
        if patient.medical_history:
            story.append(Paragraph("Medical History", styles['Heading2']))
            story.append(Paragraph(patient.medical_history, styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Allergies
        if patient.allergies:
            story.append(Paragraph("Allergies", styles['Heading2']))
            story.append(Paragraph(patient.allergies, styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Diseases
        if patient.patient_diseases.exists():
            story.append(Paragraph("Diseases", styles['Heading2']))
            disease_data = [['Disease', 'Category', 'Diagnosed Date', 'Status']]
            for pd in patient.patient_diseases.all():
                disease_data.append([
                    pd.disease.name,
                    pd.disease.get_category_display(),
                    pd.diagnosed_date.strftime('%Y-%m-%d'),
                    pd.get_status_display()
                ])
            
            disease_table = Table(disease_data)
            disease_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(disease_table)
            story.append(Spacer(1, 20))
        
        # Recent Tests
        recent_tests = patient.tests.order_by('-test_date')[:10]
        if recent_tests.exists():
            story.append(Paragraph("Recent Tests", styles['Heading2']))
            test_data = [['Test Name', 'Type', 'Date', 'Status', 'Results']]
            for test in recent_tests:
                test_data.append([
                    test.test_name,
                    test.get_test_type_display(),
                    test.test_date.strftime('%Y-%m-%d'),
                    test.get_status_display(),
                    test.results[:50] + '...' if len(test.results) > 50 else test.results
                ])
            
            test_table = Table(test_data)
            test_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(test_table)
            story.append(Spacer(1, 20))
        
        # Recent Treatments
        recent_treatments = patient.treatments.order_by('-start_date')[:5]
        if recent_treatments.exists():
            story.append(Paragraph("Recent Treatments", styles['Heading2']))
            treatment_data = [['Treatment', 'Disease', 'Start Date', 'End Date', 'Status']]
            for treatment in recent_treatments:
                treatment_data.append([
                    treatment.treatment_name,
                    treatment.disease.name,
                    treatment.start_date.strftime('%Y-%m-%d'),
                    treatment.end_date.strftime('%Y-%m-%d') if treatment.end_date else 'Ongoing',
                    treatment.get_status_display()
                ])
            
            treatment_table = Table(treatment_data)
            treatment_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(treatment_table)
            story.append(Spacer(1, 20))
        
        # Recent Surgeries
        recent_surgeries = patient.surgeries.order_by('-scheduled_date')[:5]
        if recent_surgeries.exists():
            story.append(Paragraph("Recent Surgeries", styles['Heading2']))
            surgery_data = [['Surgery', 'Surgeon', 'Scheduled Date', 'Status', 'Complications']]
            for surgery in recent_surgeries:
                surgery_data.append([
                    surgery.surgery_name,
                    surgery.surgeon_name,
                    surgery.scheduled_date.strftime('%Y-%m-%d'),
                    surgery.get_status_display(),
                    surgery.get_complications_display()
                ])
            
            surgery_table = Table(surgery_data)
            surgery_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(surgery_table)
        
        # Build PDF
        doc.build(story)
        
        # Update report
        report.status = 'COMPLETED'
        report.file_path = filepath
        report.completed_at = timezone.now()
        report.save()
        
        return f"Patient record PDF generated successfully: {filepath}"
        
    except Exception as e:
        report = Report.objects.get(id=report_id)
        report.status = 'FAILED'
        report.save()
        raise e


@shared_task
def generate_test_results_pdf(report_id):
    """Generate test results PDF report"""
    try:
        report = Report.objects.get(id=report_id)
        report.status = 'GENERATING'
        report.save()
        
        patient_id = report.parameters.get('patient_id')
        test_ids = report.parameters.get('test_ids', [])
        
        patient = Patient.objects.select_related('user', 'doctor__user').get(id=patient_id)
        
        if test_ids:
            tests = Test.objects.filter(id__in=test_ids, patient=patient)
        else:
            tests = Test.objects.filter(patient=patient).order_by('-test_date')
        
        # Create PDF
        filename = f"test_results_{patient.patient_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("Test Results Report", title_style))
        story.append(Spacer(1, 20))
        
        # Patient Information
        patient_info = [
            ['Patient ID:', patient.patient_id],
            ['Name:', patient.user.get_full_name()],
            ['Doctor:', patient.doctor.user.get_full_name()],
            ['Report Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        patient_table = Table(patient_info, colWidths=[2*inch, 4*inch])
        patient_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ]))
        
        story.append(patient_table)
        story.append(Spacer(1, 20))
        
        # Test Results
        for test in tests:
            story.append(Paragraph(f"Test: {test.test_name}", styles['Heading2']))
            
            test_data = [
                ['Test Type:', test.get_test_type_display()],
                ['Disease:', test.disease.name],
                ['Test Date:', test.test_date.strftime('%Y-%m-%d %H:%M')],
                ['Status:', test.get_status_display()],
                ['Normal Range:', test.normal_range or 'Not specified'],
            ]
            
            test_table = Table(test_data, colWidths=[2*inch, 4*inch])
            test_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ]))
            
            story.append(test_table)
            
            if test.results:
                story.append(Paragraph("Results:", styles['Heading3']))
                story.append(Paragraph(test.results, styles['Normal']))
            
            if test.notes:
                story.append(Paragraph("Notes:", styles['Heading3']))
                story.append(Paragraph(test.notes, styles['Normal']))
            
            story.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(story)
        
        # Update report
        report.status = 'COMPLETED'
        report.file_path = filepath
        report.completed_at = timezone.now()
        report.save()
        
        return f"Test results PDF generated successfully: {filepath}"
        
    except Exception as e:
        report = Report.objects.get(id=report_id)
        report.status = 'FAILED'
        report.save()
        raise e


@shared_task
def generate_treatment_summary_pdf(report_id):
    """Generate treatment summary PDF report"""
    try:
        report = Report.objects.get(id=report_id)
        report.status = 'GENERATING'
        report.save()
        
        patient_id = report.parameters.get('patient_id')
        treatment_ids = report.parameters.get('treatment_ids', [])
        
        patient = Patient.objects.select_related('user', 'doctor__user').get(id=patient_id)
        
        if treatment_ids:
            treatments = Treatment.objects.filter(id__in=treatment_ids, patient=patient).prefetch_related('treatment_medicines__medicine')
        else:
            treatments = Treatment.objects.filter(patient=patient).prefetch_related('treatment_medicines__medicine').order_by('-start_date')
        
        # Create PDF
        filename = f"treatment_summary_{patient.patient_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("Treatment Summary Report", title_style))
        story.append(Spacer(1, 20))
        
        # Patient Information
        patient_info = [
            ['Patient ID:', patient.patient_id],
            ['Name:', patient.user.get_full_name()],
            ['Doctor:', patient.doctor.user.get_full_name()],
            ['Report Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        patient_table = Table(patient_info, colWidths=[2*inch, 4*inch])
        patient_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ]))
        
        story.append(patient_table)
        story.append(Spacer(1, 20))
        
        # Treatments
        for treatment in treatments:
            story.append(Paragraph(f"Treatment: {treatment.treatment_name}", styles['Heading2']))
            
            treatment_data = [
                ['Disease:', treatment.disease.name],
                ['Start Date:', treatment.start_date.strftime('%Y-%m-%d')],
                ['End Date:', treatment.end_date.strftime('%Y-%m-%d') if treatment.end_date else 'Ongoing'],
                ['Status:', treatment.get_status_display()],
            ]
            
            treatment_table = Table(treatment_data, colWidths=[2*inch, 4*inch])
            treatment_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ]))
            
            story.append(treatment_table)
            
            if treatment.description:
                story.append(Paragraph("Description:", styles['Heading3']))
                story.append(Paragraph(treatment.description, styles['Normal']))
            
            # Medicines
            if treatment.treatment_medicines.exists():
                story.append(Paragraph("Medicines:", styles['Heading3']))
                medicine_data = [['Medicine', 'Dosage', 'Frequency', 'Duration']]
                for tm in treatment.treatment_medicines.all():
                    medicine_data.append([
                        tm.medicine.name,
                        tm.dosage,
                        tm.frequency,
                        f"{tm.duration_days} days"
                    ])
                
                medicine_table = Table(medicine_data)
                medicine_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(medicine_table)
            
            if treatment.notes:
                story.append(Paragraph("Notes:", styles['Heading3']))
                story.append(Paragraph(treatment.notes, styles['Normal']))
            
            story.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(story)
        
        # Update report
        report.status = 'COMPLETED'
        report.file_path = filepath
        report.completed_at = timezone.now()
        report.save()
        
        return f"Treatment summary PDF generated successfully: {filepath}"
        
    except Exception as e:
        report = Report.objects.get(id=report_id)
        report.status = 'FAILED'
        report.save()
        raise e


@shared_task
def generate_surgery_report_pdf(report_id):
    """Generate surgery report PDF"""
    try:
        report = Report.objects.get(id=report_id)
        report.status = 'GENERATING'
        report.save()
        
        surgery_id = report.parameters.get('surgery_id')
        surgery = Surgery.objects.select_related('patient__user', 'patient__doctor__user').get(id=surgery_id)
        
        # Create PDF
        filename = f"surgery_report_{surgery.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("Surgery Report", title_style))
        story.append(Spacer(1, 20))
        
        # Surgery Information
        surgery_data = [
            ['Surgery Name:', surgery.surgery_name],
            ['Patient ID:', surgery.patient.patient_id],
            ['Patient Name:', surgery.patient.user.get_full_name()],
            ['Surgeon:', surgery.surgeon_name],
            ['Scheduled Date:', surgery.scheduled_date.strftime('%Y-%m-%d %H:%M')],
            ['Actual Date:', surgery.actual_date.strftime('%Y-%m-%d %H:%M') if surgery.actual_date else 'Not performed'],
            ['Status:', surgery.get_status_display()],
            ['Complications:', surgery.get_complications_display()],
        ]
        
        surgery_table = Table(surgery_data, colWidths=[2*inch, 4*inch])
        surgery_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ]))
        
        story.append(surgery_table)
        story.append(Spacer(1, 20))
        
        if surgery.description:
            story.append(Paragraph("Description:", styles['Heading2']))
            story.append(Paragraph(surgery.description, styles['Normal']))
            story.append(Spacer(1, 20))
        
        if surgery.complications_description:
            story.append(Paragraph("Complications Details:", styles['Heading2']))
            story.append(Paragraph(surgery.complications_description, styles['Normal']))
            story.append(Spacer(1, 20))
        
        if surgery.notes:
            story.append(Paragraph("Notes:", styles['Heading2']))
            story.append(Paragraph(surgery.notes, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        # Update report
        report.status = 'COMPLETED'
        report.file_path = filepath
        report.completed_at = timezone.now()
        report.save()
        
        return f"Surgery report PDF generated successfully: {filepath}"
        
    except Exception as e:
        report = Report.objects.get(id=report_id)
        report.status = 'FAILED'
        report.save()
        raise e


@shared_task
def generate_patients_per_city_excel(report_id):
    """Generate patients per city Excel report"""
    try:
        if not PANDAS_AVAILABLE:
            report = Report.objects.get(id=report_id)
            report.status = 'FAILED'
            report.save()
            return "Excel generation requires pandas and openpyxl packages"
        
        report = Report.objects.get(id=report_id)
        report.status = 'GENERATING'
        report.save()
        
        city_ids = report.parameters.get('city_ids', [])
        
        if city_ids:
            cities = City.objects.filter(id__in=city_ids)
        else:
            cities = City.objects.all()
        
        # Create Excel file
        filename = f"patients_per_city_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Patients per City"
        
        # Headers
        headers = ['City', 'State', 'Country', 'Centers Count', 'Doctors Count', 'Patients Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        row = 2
        for city in cities:
            centers_count = city.centers.count()
            doctors_count = sum(center.doctors.count() for center in city.centers.all())
            patients_count = sum(doctor.patients.count() for center in city.centers.all() for doctor in center.doctors.all())
            
            ws.cell(row=row, column=1, value=city.name)
            ws.cell(row=row, column=2, value=city.state)
            ws.cell(row=row, column=3, value=city.country)
            ws.cell(row=row, column=4, value=centers_count)
            ws.cell(row=row, column=5, value=doctors_count)
            ws.cell(row=row, column=6, value=patients_count)
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        
        # Update report
        report.status = 'COMPLETED'
        report.file_path = filepath
        report.completed_at = timezone.now()
        report.save()
        
        return f"Patients per city Excel generated successfully: {filepath}"
        
    except Exception as e:
        report = Report.objects.get(id=report_id)
        report.status = 'FAILED'
        report.save()
        raise e


@shared_task
def generate_common_diseases_excel(report_id):
    """Generate common diseases Excel report"""
    try:
        if not PANDAS_AVAILABLE:
            report = Report.objects.get(id=report_id)
            report.status = 'FAILED'
            report.save()
            return "Excel generation requires pandas and openpyxl packages"
        
        report = Report.objects.get(id=report_id)
        report.status = 'GENERATING'
        report.save()
        
        center_ids = report.parameters.get('center_ids', [])
        start_date = report.parameters.get('start_date')
        end_date = report.parameters.get('end_date')
        
        # Filter diseases
        diseases_query = Disease.objects.all()
        
        if center_ids:
            diseases_query = diseases_query.filter(
                patient_diseases__patient__doctor__center_id__in=center_ids
            ).distinct()
        
        if start_date and end_date:
            diseases_query = diseases_query.filter(
                patient_diseases__diagnosed_date__range=[start_date, end_date]
            ).distinct()
        
        # Create Excel file
        filename = f"common_diseases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Common Diseases"
        
        # Headers
        headers = ['Disease Name', 'Category', 'ICD Code', 'Patient Count', 'Centers Affected']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        row = 2
        for disease in diseases_query:
            patient_count = disease.patient_diseases.count()
            centers_affected = disease.patient_diseases.values_list(
                'patient__doctor__center__name', flat=True
            ).distinct().count()
            
            ws.cell(row=row, column=1, value=disease.name)
            ws.cell(row=row, column=2, value=disease.get_category_display())
            ws.cell(row=row, column=3, value=disease.icd_code or '')
            ws.cell(row=row, column=4, value=patient_count)
            ws.cell(row=row, column=5, value=centers_affected)
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        
        # Update report
        report.status = 'COMPLETED'
        report.file_path = filepath
        report.completed_at = timezone.now()
        report.save()
        
        return f"Common diseases Excel generated successfully: {filepath}"
        
    except Exception as e:
        report = Report.objects.get(id=report_id)
        report.status = 'FAILED'
        report.save()
        raise e
